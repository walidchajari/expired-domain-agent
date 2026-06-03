import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import settings

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

CATEGORY_FILLS = {
    "Brandable": PatternFill(start_color="E8D4F0", end_color="E8D4F0", fill_type="solid"),
    "Geo": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
    "AI": PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid"),
    "Fintech": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
    "High Commercial Keyword": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
}
ALT_ROW_FILL = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")


# ---------------------------------------------------------------------------
# Report 1: Available Domains (ALL raw data, no AI)
# ---------------------------------------------------------------------------

AVAILABLE_COLUMNS = [
    ("Domain", 28), ("LE", 5), ("BL", 10), ("DP", 10),
    ("WBY", 8), ("ABY", 8), ("ACR", 8), ("MMGR", 12),
    ("DMOZ", 8), ("REG", 5), ("C", 5), ("N", 5), ("O", 5),
    ("B", 5), ("I", 5), ("D", 5), ("AddDate", 12), ("RDT", 5),
    ("Status", 10),
]

AVAILABLE_KEY_MAP = {
    "LE": "length",
    "BL": "bl",
    "DP": "dp",
    "WBY": "wby",
    "ABY": "aby",
    "ACR": "acr",
    "MMGR": "mmgr",
    "DMOZ": "dmoz",
    "REG": "reg",
    "C": "reg_c",
    "N": "reg_n",
    "O": "reg_o",
    "B": "reg_b",
    "I": "reg_i",
    "D": "reg_d",
    "AddDate": "adddate",
    "RDT": "rdt",
    "Status": "status",
}


def _build_available_rows(domains: list[dict]) -> list[dict]:
    rows = []
    for d in domains:
        row = {"Domain": d["domain"]}
        for col_name, _ in AVAILABLE_COLUMNS:
            if col_name == "Domain":
                continue
            key = AVAILABLE_KEY_MAP.get(col_name, col_name.lower())
            row[col_name] = d.get(key, "")
        rows.append(row)
    return rows


def _build_investor_rows(rankings: list[dict]) -> list[dict]:
    rows = []
    for rank, d in enumerate(rankings, start=1):
        row = {"Rank": rank, "Domain": d["domain"]}
        eng = d.get("english_scores", {})

        row["Category"] = d.get("category", "")
        row["Final Score"] = round(d.get("final_score", 0), 2)
        row["Brandability"] = d.get("brandability", 0)
        row["Pronounceability"] = eng.get("pronounceability_score", 50) if isinstance(eng, dict) else 50
        row["Commercial Intent"] = eng.get("commercial_intent_score", 0) if isinstance(eng, dict) else 0
        row["Startup Pattern"] = d.get("startup_pattern_score", 0)
        row["Liquid Score"] = d.get("liquid_score", 0)
        row["Geo Score"] = d.get("geo_score", 0)
        row["REG"] = d.get("reg", 0)
        row["DP"] = d.get("dp", "")
        row["BL"] = d.get("bl", "")
        row["WBY"] = d.get("wby", "")
        try:
            wby = int(d.get("wby", "0")) if str(d.get("wby", "0")).isdigit() else 0
            age = datetime.now().year - wby if 1900 < wby <= datetime.now().year else 0
        except (ValueError, TypeError):
            age = 0
        row["Age"] = age
        row["Sale Prob"] = f"{d.get('probability_of_sale', 0):.0f}%"
        row["Wholesale"] = d.get("estimated_wholesale_price", "")
        row["End User"] = d.get("estimated_end_user_price", "")
        row["AI Rec"] = d.get("ai_recommendation", "")
        row["Conf."] = f"{d.get('ai_confidence', 0):.0f}%" if d.get("ai_confidence") else ""
        row["Reason"] = d.get("reason_for_selection", "")

        rows.append(row)
    return rows


# ---------------------------------------------------------------------------
# Report 2: Investor Domains (scored ≥ 70)
# ---------------------------------------------------------------------------

INVESTOR_COLUMNS = [
    ("Rank", 6), ("Domain", 28), ("Category", 18),
    ("Final Score", 10), ("Brandability", 11),
    ("Pronounceability", 14), ("Commercial Intent", 16),
    ("Startup Pattern", 14), ("Liquid Score", 12),
    ("Geo Score", 10), ("REG", 5), ("DP", 7),
    ("BL", 7), ("WBY", 6), ("Age", 5),
    ("Wholesale", 18), ("End User", 18),
    ("Sale Prob", 10), ("AI Rec", 8), ("Conf.", 7),
    ("Reason", 55),
]


def _apply_formatting_available(filepath: Path) -> None:
    wb = load_workbook(filepath)
    ws = wb.active

    for col_idx, (col_name, width) in enumerate(AVAILABLE_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width

    ws.row_dimensions[1].height = 28
    for col_idx in range(1, len(AVAILABLE_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(AVAILABLE_COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Calibri", size=11)
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

        domain_cell = ws.cell(row=row_idx, column=1)
        domain_cell.alignment = Alignment(horizontal="left", vertical="center")
        domain_cell.font = Font(name="Calibri", size=11, bold=True)

    last_col = get_column_letter(len(AVAILABLE_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"
    wb.save(filepath)


def _apply_formatting_investor(filepath: Path) -> None:
    wb = load_workbook(filepath)
    ws = wb.active
    columns = INVESTOR_COLUMNS

    col_letters = {}
    for col_idx, (col_name, width) in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        col_letters[col_name] = letter
        ws.column_dimensions[letter].width = width

    ws.row_dimensions[1].height = 28
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"

    category_col = col_letters.get("Category", "C")
    domain_col = col_letters.get("Domain", "B")
    reason_col = col_letters.get("Reason", get_column_letter(len(columns)))
    ai_rec_col = col_letters.get("AI Rec", "")
    score_col = col_letters.get("Final Score", "D")

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Calibri", size=11)
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

        domain_cell = ws[f"{domain_col}{row_idx}"]
        domain_cell.alignment = Alignment(horizontal="left", vertical="center")
        domain_cell.font = Font(name="Calibri", size=11, bold=True)

        cat_cell = ws[f"{category_col}{row_idx}"]
        cat_val = str(cat_cell.value or "")
        if cat_val in CATEGORY_FILLS:
            cat_cell.fill = CATEGORY_FILLS[cat_val]
            cat_cell.font = Font(name="Calibri", size=11, bold=True)

        reason_cell = ws[f"{reason_col}{row_idx}"]
        reason_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row_idx].height = 28

        score_cell = ws[f"{score_col}{row_idx}"]
        try:
            score = float(score_cell.value or 0)
        except (ValueError, TypeError):
            score = 0
        if score > 85:
            score_cell.fill = GREEN_FILL
            score_cell.font = Font(name="Calibri", size=11, bold=True)
        elif score >= 70:
            score_cell.fill = YELLOW_FILL
            score_cell.font = Font(name="Calibri", size=11, bold=True)
        else:
            score_cell.fill = RED_FILL
            score_cell.font = Font(name="Calibri", size=11)

        if ai_rec_col:
            ai_cell = ws[f"{ai_rec_col}{row_idx}"]
            ai_val = str(ai_cell.value or "")
            if ai_val == "BUY":
                ai_cell.font = Font(name="Calibri", size=11, bold=True, color="006100")
            elif ai_val == "MAYBE":
                ai_cell.font = Font(name="Calibri", size=11, bold=True, color="9C6500")
            elif ai_val == "PASS":
                ai_cell.font = Font(name="Calibri", size=11, color="9C0006")

        for price_col_name in ("Wholesale", "End User"):
            if price_col_name in col_letters:
                price_cell = ws[f"{col_letters[price_col_name]}{row_idx}"]
                price_cell.font = Font(name="Calibri", size=11, color="2E7D32")

    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"
    wb.save(filepath)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_available_report(domains: list[dict]) -> Path:
    today = datetime.now().strftime("%Y_%m_%d")
    filename = f"available_domains_{today}.xlsx"
    filepath = settings.excel_dir / filename

    rows = _build_available_rows(domains)
    columns = [c[0] for c in AVAILABLE_COLUMNS]
    df = pd.DataFrame(rows, columns=columns)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Available Domains", index=False)

    _apply_formatting_available(filepath)
    logger.info("Available report generated: %s (%d domains)", filepath, len(domains))
    return filepath


def generate_investor_report(rankings: list[dict]) -> Path:
    today = datetime.now().strftime("%Y_%m_%d")
    filename = f"investor_domains_{today}.xlsx"
    filepath = settings.excel_dir / filename

    rows = _build_investor_rows(rankings)
    columns = [c[0] for c in INVESTOR_COLUMNS]
    df = pd.DataFrame(rows, columns=columns)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Investor Domains", index=False)

    _apply_formatting_investor(filepath)
    logger.info("Investor report generated: %s (%d domains)", filepath, len(rankings))
    return filepath


def generate_both_reports(all_domains: list[dict], investor_rankings: list[dict]) -> list[Path]:
    paths = []
    paths.append(generate_available_report(all_domains))
    paths.append(generate_investor_report(investor_rankings))
    return paths


def export_summary_text(investor_domains: list[dict]) -> str:
    if not investor_domains:
        return "No investor-grade domains found today."
    lines = [
        f"Total domains analyzed: {len(investor_domains)}",
        f"Top Score: {investor_domains[0].get('final_score', 0):.1f}",
        f"Best Domain: {investor_domains[0]['domain']}",
        f"Category: {investor_domains[0].get('category', 'N/A')}",
        "",
        "Top 5:",
    ]
    for i, d in enumerate(investor_domains[:5], 1):
        cat = d.get("category", "")
        extra = f" [{cat}]" if cat else ""
        lines.append(f"  {i}. {d['domain']} ({d.get('final_score', 0):.1f}){extra}")
    return "\n".join(lines)